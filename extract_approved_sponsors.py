from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader


DEFAULT_PDF = Path(
    r"D:\My Files\hw\AUS Govt Docs, Visa, COE, OSHC\PR homework\Approved_Sponsors.pdf"
)
DEFAULT_OUTPUT = Path(
    r"D:\My Files\hw\Portfolio Projects\Projects\PR homework\approved_sponsors.xlsx"
)


NOISE_EXACT = {
    "[please select protective marking from the home tab]",
    "released by department of home affairs",
    "under the freedom of information act 1982",
}


def normalize_line(raw_line: str) -> str:
    """Normalize PDF text extraction whitespace without changing company names."""
    return re.sub(r"\s+", " ", raw_line.replace("\u00a0", " ")).strip()


def is_noise_line(line: str) -> bool:
    """Return True for watermark/header/footer lines, not sponsor names."""
    if not line:
        return True

    lower = line.lower()

    if lower in NOISE_EXACT:
        return True
    if re.fullmatch(r"\d+", line):
        return True
    if "please select protective marking" in lower:
        return True
    if lower.startswith("freedom of information request"):
        return True
    if "accredited sponsors" in lower and "standard business sponsorship" in lower:
        return True

    return False


def extract_sponsors(pdf_path: Path) -> list[dict[str, object]]:
    reader = PdfReader(str(pdf_path))
    records: list[dict[str, object]] = []
    stop_after_notes = False

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for raw_line in text.splitlines():
            line = normalize_line(raw_line)
            lower = line.lower()

            if lower.startswith("notes:") or lower.startswith("caveats:"):
                stop_after_notes = True
                break

            if is_noise_line(line):
                continue

            records.append(
                {
                    "Source Order": len(records) + 1,
                    "Sponsor Name": line,
                    "PDF Page": page_number,
                }
            )

        if stop_after_notes:
            break

    return records


BARE_SUFFIXES = {
    "inc",
    "ltd",
    "pty ltd",
    "pty limited",
    "limited",
    "llc",
    "co",
    "corp",
    "corporation",
    "pl",
    "p/l",
    "inc.",
    "ltd.",
    "co.",
    "corp.",
}


def merge_wrapped_suffixes(
    records: list[dict[str, object]]
) -> list[dict[str, object]]:
    """Merge a bare corporate-suffix line (e.g. 'INC') into the previous
    record when the PDF has wrapped a long company name across two lines.
    """
    merged: list[dict[str, object]] = []

    for record in records:
        name = str(record["Sponsor Name"]).strip()
        lower = name.casefold()

        if (
            merged
            and lower in BARE_SUFFIXES
            and merged[-1]["PDF Page"] == record["PDF Page"]
        ):
            prev_name = str(merged[-1]["Sponsor Name"]).strip()
            if not prev_name.casefold().rstrip(".").endswith(
                tuple(s.rstrip(".") for s in BARE_SUFFIXES)
            ):
                merged[-1]["Sponsor Name"] = f"{prev_name} {name}"
                continue

        merged.append(dict(record))

    for index, record in enumerate(merged, start=1):
        record["Source Order"] = index

    return merged


def unique_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    seen: set[str] = set()
    output: list[dict[str, object]] = []

    for record in records:
        key = str(record["Sponsor Name"]).casefold()
        if key in seen:
            continue
        seen.add(key)
        output.append(
            {
                "Source Order": len(output) + 1,
                "Sponsor Name": record["Sponsor Name"],
                "PDF Page": record["PDF Page"],
            }
        )

    return output


def write_excel(records: list[dict[str, object]], output_path: Path) -> None:
    if not records:
        raise ValueError("No sponsor names were extracted from the PDF.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Approved Sponsors"

    headers = ["Source Order", "Sponsor Name", "PDF Page"]
    sheet.append(headers)
    for record in records:
        sheet.append([record["Source Order"], record["Sponsor Name"], record["PDF Page"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="right")
        row[1].alignment = Alignment(wrap_text=False)
        row[2].alignment = Alignment(horizontal="right")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    table = Table(displayName="ApprovedSponsors", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    widths = {"A": 14, "B": 72, "C": 10}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 18

    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract approved sponsor names from the Home Affairs PDF into Excel."
    )
    parser.add_argument("--pdf", type=Path, default=DEFAULT_PDF, help="Input PDF path.")
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_OUTPUT, help="Output .xlsx path."
    )
    parser.add_argument(
        "--unique",
        action="store_true",
        help="Keep only the first occurrence of each sponsor name.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.pdf.exists():
        raise FileNotFoundError(f"Input PDF was not found: {args.pdf}")

    records = extract_sponsors(args.pdf)
    records = merge_wrapped_suffixes(records)
    if args.unique:
        records = unique_records(records)
    write_excel(records, args.output)
    print(f"Extracted {len(records):,} sponsor names")
    print(f"Saved Excel file to: {args.output}")


if __name__ == "__main__":
    main()
