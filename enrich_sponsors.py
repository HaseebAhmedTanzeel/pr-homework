"""
Phase 2: enrich approved-sponsor company names with website, Australian state,
industry, and a confidence rating based on independent source agreement.

REVISION NOTE (v2): the first version of this script used Gemini for every
company. In practice the free-tier daily quota turned out to be 20
requests/day (not the 1,500/day the docs implied) - completely unworkable at
3,389 companies. v2 removes that bottleneck: extraction is done with
deterministic rules directly from search results by default, so there is no
API quota at all. Gemini becomes optional (--use-llm) as a slow bonus pass
you can run later just on the leftover unresolved rows if you want to.

Pipeline (free, no ABN/ABR lookups):
  Tier 0 - Wikidata: official, structured, zero-cost lookup for companies that
           already have a Wikidata entry (mostly large/well-known corporates).
  Tier 1 - DuckDuckGo web search (unofficial, free) -> results are filtered
           against a blacklist of directories/social sites, and the top
           remaining result is taken as the likely official site. Confidence
           comes from (a) whether the search engine ranked it first among
           real candidates and (b) whether another independent result
           separately mentions the same domain.
  Tier 2 - Unresolved: no usable/agreeing evidence found. Recorded honestly
           rather than guessed.
  Optional bonus pass (--use-llm): re-examines only rows still Low/None after
           Tier 1, using a free-tier Gemini call per row. Capped hard at the
           real observed daily quota so it can't crash the whole run.

Checkpointing: every company's result is committed to a local SQLite database
the moment it's produced, so the script can be killed (Ctrl+C, power loss,
network blip) at any time and simply resumed later with the same command -
already-processed companies are skipped automatically.

Setup (one-time):
    pip install ddgs openpyxl
    (only needed if you use --use-llm:  pip install google-genai
     and get a free key from https://aistudio.google.com/apikey)

Usage:
    python enrich_sponsors.py --input approved_sponsors_unique.xlsx --output enriched_sponsors.xlsx
    (re-run the same command any time to resume from where it stopped)

    Optional flags:
    --limit 50          process only the first 50 unprocessed companies (good for a test run)
    --db enrichment.db  custom checkpoint database path (default: enrichment.db)
    --skip-wikidata     skip Tier 0 and go straight to search-based extraction
    --use-llm           add the optional Gemini bonus pass on Low/None rows
                         (requires GEMINI_API_KEY env var; capped at 20/day)
"""

from __future__ import annotations

import argparse
import difflib
import json
import os
import random
import re
import sqlite3
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MAX_SEARCH_RESULTS = 6
DDG_MIN_DELAY_SECONDS = 2.0

GEMINI_MODEL = "gemini-2.5-flash-lite"
GEMINI_MIN_DELAY_SECONDS = 4.5
# Observed empirically from a real run's 429 error - do not trust published
# docs here, they were wrong. Update this if your own account's error message
# reports a different number.
GEMINI_FREE_RPD = 20

MAX_CONSECUTIVE_ERRORS = 8

AUS_STATES = ["NSW", "VIC", "QLD", "WA", "SA", "TAS", "ACT", "NT"]

STATE_KEYWORDS = {
    "NSW": ["new south wales", "nsw", "sydney", "newcastle", "wollongong"],
    "VIC": ["victoria", "vic", "melbourne", "geelong", "ballarat"],
    "QLD": ["queensland", "qld", "brisbane", "gold coast", "townsville", "cairns"],
    "WA": ["western australia", " wa ", "perth", "fremantle"],
    "SA": ["south australia", " sa ", "adelaide"],
    "TAS": ["tasmania", "tas", "hobart", "launceston"],
    "ACT": ["australian capital territory", "canberra", " act "],
    "NT": ["northern territory", "darwin", " nt "],
}

INDUSTRY_KEYWORDS = {
    "Agriculture, Forestry and Fishing": ["farm", "agricult", "forestry", "fishing", "livestock", "crop"],
    "Mining": ["mining", "mine ", "mineral", "coal", "resources ltd", "exploration"],
    "Manufacturing": ["manufactur", "factory", "production plant", "fabrication"],
    "Electricity, Gas, Water and Waste Services": ["electricity", "energy supply", "gas supply", "water utility", "waste management", "recycling"],
    "Construction": ["construction", "builders", "building contractor", "civil works", "engineering construction"],
    "Wholesale Trade": ["wholesale", "distributor", "supplier of"],
    "Retail Trade": ["retail", "store", "supermarket", "shop "],
    "Accommodation and Food Services": ["hotel", "restaurant", "cafe", "accommodation", "catering", "hospitality"],
    "Transport, Postal and Warehousing": ["transport", "logistics", "freight", "shipping", "warehousing", "courier"],
    "Information Media and Telecommunications": ["software", "telecommunicat", "media company", "broadcasting", "IT services", "technology solutions"],
    "Financial and Insurance Services": ["bank", "banking", "insurance", "financial services", "superannuation", "investment fund"],
    "Rental, Hiring and Real Estate Services": ["real estate", "property management", "leasing", "rental services"],
    "Professional, Scientific and Technical Services": ["consulting", "engineering services", "law firm", "legal services", "accounting firm", "scientific research", "architecture"],
    "Administrative and Support Services": ["administrative services", "labour hire", "recruitment", "cleaning services", "security services"],
    "Public Administration and Safety": ["government department", "council", "local government", "public service", "defence force", "police"],
    "Education and Training": ["school", "college", "university", "tafe", "training provider", "education"],
    "Health Care and Social Assistance": ["hospital", "health care", "medical centre", "aged care", "disability services", "clinic"],
    "Arts and Recreation Services": ["museum", "gallery", "theatre", "sporting club", "recreation centre", "arts organisation"],
}

DIRECTORY_DOMAINS = {
    "facebook.com", "linkedin.com", "instagram.com", "twitter.com", "x.com",
    "youtube.com", "tiktok.com", "yellowpages.com.au", "truelocal.com.au",
    "whitepages.com.au", "hotfrog.com.au", "startlocal.com.au", "yelp.com",
    "seek.com.au", "indeed.com", "glassdoor.com", "glassdoor.com.au",
    "crunchbase.com", "bloomberg.com", "dnb.com", "opencorporates.com",
    "zoominfo.com", "rocketreach.co", "apollo.io", "wikipedia.org",
    "wikidata.org", "google.com", "abr.business.gov.au", "abn.business.gov.au",
    "informdirect.com.au", "quickbusinesssearch.com.au", "cluey.com.au",
}


# ---------------------------------------------------------------------------
# SQLite checkpoint store
# ---------------------------------------------------------------------------

def init_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS enrichment (
            source_order INTEGER PRIMARY KEY,
            sponsor_name TEXT NOT NULL,
            website TEXT,
            state TEXT,
            industry TEXT,
            confidence TEXT,
            confidence_reason TEXT,
            resolved_by TEXT,
            status TEXT NOT NULL,
            raw_response TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.commit()
    return conn


def already_done(conn: sqlite3.Connection, source_order: int) -> bool:
    row = conn.execute(
        "SELECT status FROM enrichment WHERE source_order = ?", (source_order,)
    ).fetchone()
    return row is not None and row[0] in ("done", "unresolved")


def save_result(conn: sqlite3.Connection, source_order: int, name: str, result: dict) -> None:
    conn.execute(
        """
        INSERT INTO enrichment
            (source_order, sponsor_name, website, state, industry,
             confidence, confidence_reason, resolved_by, status, raw_response)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(source_order) DO UPDATE SET
            website=excluded.website, state=excluded.state, industry=excluded.industry,
            confidence=excluded.confidence, confidence_reason=excluded.confidence_reason,
            resolved_by=excluded.resolved_by, status=excluded.status,
            raw_response=excluded.raw_response, updated_at=CURRENT_TIMESTAMP
        """,
        (
            source_order,
            name,
            result.get("website"),
            result.get("state"),
            result.get("industry"),
            result.get("confidence"),
            result.get("confidence_reason"),
            result.get("resolved_by"),
            result.get("status", "done"),
            json.dumps(result.get("raw_response"))[:5000] if result.get("raw_response") else None,
        ),
    )
    conn.commit()  # commit immediately -> this IS the checkpoint


# ---------------------------------------------------------------------------
# Tier 0: Wikidata (official, free, structured)
# ---------------------------------------------------------------------------

def wikidata_lookup(name: str) -> dict | None:
    """Look up a company on Wikidata. Returns a result dict if a confident
    match with an official website is found, else None."""
    try:
        search_url = (
            "https://www.wikidata.org/w/api.php?action=wbsearchentities"
            f"&search={urllib.parse.quote(name)}&language=en&format=json&limit=1&type=item"
        )
        req = urllib.request.Request(search_url, headers={"User-Agent": "SponsorEnrichment/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        hits = data.get("search", [])
        if not hits:
            return None

        qid = hits[0]["id"]
        entity_url = f"https://www.wikidata.org/wiki/Special:EntityData/{qid}.json"
        req2 = urllib.request.Request(entity_url, headers={"User-Agent": "SponsorEnrichment/1.0"})
        with urllib.request.urlopen(req2, timeout=15) as resp2:
            entity_data = json.loads(resp2.read().decode("utf-8"))

        claims = entity_data["entities"][qid]["claims"]

        country_claims = claims.get("P17", [])
        is_australian = any(
            c["mainsnak"]["datavalue"]["value"]["id"] == "Q408"
            for c in country_claims
            if c["mainsnak"].get("datavalue")
        )
        if not is_australian:
            return None

        website = None
        if "P856" in claims:
            website = claims["P856"][0]["mainsnak"]["datavalue"]["value"]

        if not website:
            return None

        return {
            "website": website,
            "state": "Unknown",
            "industry": "Unknown",
            "confidence": "High",
            "confidence_reason": "Matched an Australian entity on Wikidata with an official website property (P856).",
            "resolved_by": "wikidata",
            "status": "done",
            "raw_response": {"qid": qid},
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Tier 1: DuckDuckGo search + rule-based extraction (no LLM, no quota)
# ---------------------------------------------------------------------------

def ddg_search(query: str, max_results: int = MAX_SEARCH_RESULTS) -> list[dict]:
    from ddgs import DDGS

    with DDGS() as ddgs:
        results = list(ddgs.text(query, max_results=max_results))
    return [
        {"title": r.get("title", ""), "href": r.get("href", ""), "body": r.get("body", "")}
        for r in results
    ]


def domain_of(url: str) -> str:
    try:
        netloc = urllib.parse.urlparse(url).netloc.lower()
        return re.sub(r"^www\.", "", netloc)
    except Exception:
        return ""


def clean_name(name: str) -> str:
    n = name.lower()
    n = re.sub(r"\((.*?)\)", " ", n)  # drop parenthetical qualifiers e.g. "(Australia)"
    n = re.sub(
        r"\b(pty\.?\s*ltd\.?|pty\.?\s*limited|limited|ltd\.?|incorporated|inc\.?|holdings?|group)\b",
        " ", n,
    )
    n = re.sub(r"[^a-z0-9 ]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def detect_state(combined_text: str) -> str:
    scores = {}
    for state, keywords in STATE_KEYWORDS.items():
        scores[state] = sum(combined_text.count(kw) for kw in keywords)
    best_state = max(scores, key=scores.get)
    return best_state if scores[best_state] > 0 else "Unknown"


def detect_industry(combined_text: str) -> str:
    scores = {}
    for industry, keywords in INDUSTRY_KEYWORDS.items():
        scores[industry] = sum(combined_text.count(kw) for kw in keywords)
    best_industry = max(scores, key=scores.get)
    return best_industry if scores[best_industry] > 0 else "Unknown"


def heuristic_extract(name: str, results: list[dict]) -> dict:
    """Pick the likely official site using DuckDuckGo's own relevance ranking
    (a real signal - don't discard it in favour of naive string matching,
    which fails badly on acronym-style domains like 'ems.com.au') plus a
    check for independent corroboration."""

    candidates = [r for r in results if domain_of(r["href"]) not in DIRECTORY_DOMAINS]

    combined_text = " ".join(f"{r.get('title', '')} {r.get('body', '')}" for r in results).lower()
    state = detect_state(combined_text)
    industry = detect_industry(combined_text)

    if not candidates:
        return {
            "website": None, "state": state, "industry": industry,
            "confidence": "None", "confidence_reason": "No non-directory search results found.",
            "resolved_by": "search-heuristic", "status": "unresolved",
            "raw_response": {"results": results},
        }

    top = candidates[0]
    target_domain = domain_of(top["href"])

    corroborated_by = [
        r for r in results
        if r is not top and target_domain and target_domain in f"{r.get('title','')} {r.get('body','')}".lower()
    ]

    cleaned = clean_name(name).replace(" ", "")
    base = target_domain.split(".")[0]
    similarity = difflib.SequenceMatcher(None, cleaned, base).ratio() if cleaned and base else 0.0

    if corroborated_by:
        confidence = "High"
        reason = (
            f"DuckDuckGo's top non-directory result was {target_domain}, and "
            f"{len(corroborated_by)} separate result(s) independently mention it too."
        )
        status = "done"
    elif similarity >= 0.4 or len(candidates) == 1:
        confidence = "Medium"
        reason = f"DuckDuckGo's top non-directory result was {target_domain}, not independently corroborated elsewhere in these results."
        status = "done"
    else:
        confidence = "Low"
        reason = (
            f"Top non-directory result was {target_domain}, but it doesn't closely "
            "resemble the company name and isn't corroborated - verify manually."
        )
        status = "unresolved"

    return {
        "website": target_domain,
        "state": state,
        "industry": industry,
        "confidence": confidence,
        "confidence_reason": reason,
        "resolved_by": "search-heuristic",
        "status": status,
        "raw_response": {"results": results, "similarity": round(similarity, 2)},
    }


def enrich_via_search(name: str) -> dict:
    query = f"{name} Australia"  # NOTE: no exact-phrase quotes - that was the v1 bug
    results = ddg_search(query)

    if not results:
        cleaned = clean_name(name)
        if cleaned and cleaned.lower() != name.lower():
            time.sleep(DDG_MIN_DELAY_SECONDS)
            results = ddg_search(f"{cleaned} Australia")

    if not results:
        return {
            "website": None, "state": "Unknown", "industry": "Unknown",
            "confidence": "None", "confidence_reason": "No search results returned for this company.",
            "resolved_by": "search-heuristic", "status": "unresolved",
            "raw_response": None,
        }

    return heuristic_extract(name, results)


# ---------------------------------------------------------------------------
# Optional bonus pass: Gemini LLM re-check for Low/None rows only
# ---------------------------------------------------------------------------

INDUSTRY_DIVISIONS = list(INDUSTRY_KEYWORDS.keys()) + ["Other Services", "Unknown"]


def build_llm_prompt(name: str, results: list[dict]) -> str:
    snippets_text = "\n\n".join(
        f"[Result {i + 1}] {r['title']}\nURL: {r['href']}\n{r['body']}"
        for i, r in enumerate(results)
    )
    states_list = ", ".join(AUS_STATES + ["Unknown"])
    industries_list = "\n".join(f"- {i}" for i in INDUSTRY_DIVISIONS)
    return f"""You are extracting factual data about an Australian company from search results only.
Company name: "{name}"

Search results:
{snippets_text}

Rules:
- Use ONLY the information present in the search results above.
- If unclear, say so - do not guess.
- "website" must be the company's own official domain (not a directory/social/news site), or null.
- "state" must be one of: {states_list}.
- "industry" must be exactly one of:
{industries_list}

Respond with ONLY this JSON object:
{{"website": "example.com.au or null", "state": "...", "industry": "...", "reasoning": "one short sentence"}}"""


def call_gemini(client, prompt: str) -> dict:
    from google.genai import types

    response = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(
            temperature=0.0, max_output_tokens=300, response_mime_type="application/json",
        ),
    )
    text = re.sub(r"^```json|```$", "", response.text.strip(), flags=re.MULTILINE).strip()
    return json.loads(text)


def llm_bonus_pass(conn: sqlite3.Connection, client, daily_cap: int) -> None:
    rows = conn.execute(
        "SELECT source_order, sponsor_name, raw_response FROM enrichment "
        "WHERE confidence IN ('Low', 'None') ORDER BY source_order"
    ).fetchall()

    print(f"\nBonus LLM pass: {len(rows):,} rows currently Low/None confidence. "
          f"Processing up to {daily_cap} today (real observed free-tier daily cap).")

    calls_made = 0
    for source_order, name, raw_response_json in rows:
        if calls_made >= daily_cap:
            print(f"Hit the {daily_cap}/day cap - re-run with --use-llm again tomorrow to continue.")
            break
        try:
            raw = json.loads(raw_response_json) if raw_response_json else {}
            results = raw.get("results") or []
            if not results:
                continue

            prompt = build_llm_prompt(name, results)
            extracted = call_gemini(client, prompt)
            calls_made += 1
            time.sleep(GEMINI_MIN_DELAY_SECONDS + random.uniform(0, 1.5))

            website = extracted.get("website")
            if isinstance(website, str) and website.strip().lower() == "null":
                website = None
            status = "done" if website else "unresolved"
            confidence = "Medium" if website else "None"

            save_result(conn, source_order, name, {
                "website": website,
                "state": extracted.get("state", "Unknown"),
                "industry": extracted.get("industry", "Unknown"),
                "confidence": confidence,
                "confidence_reason": "Re-evaluated by Gemini bonus pass: " + str(extracted.get("reasoning", "")),
                "resolved_by": "search+llm-bonus",
                "status": status,
                "raw_response": extracted,
            })
            print(f"[{source_order}] {name[:60]:<60} -> {confidence} (bonus pass)")

        except Exception as exc:
            print(f"[{source_order}] {name[:60]:<60} -> bonus pass ERROR: {exc}")
            if "429" in str(exc) or "quota" in str(exc).lower():
                print("Daily quota hit - re-run with --use-llm tomorrow to continue.")
                break


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def load_companies(input_path: Path) -> list[tuple[int, str]]:
    wb = load_workbook(input_path, read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    return [(row[0], row[1]) for row in rows if row[1]]


def export_to_excel(db_path: Path, output_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        """
        SELECT source_order, sponsor_name, website, state, industry,
               confidence, confidence_reason, resolved_by, status
        FROM enrichment ORDER BY source_order
        """
    ).fetchall()
    conn.close()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Enriched Sponsors"
    sheet.append([
        "Source Order", "Sponsor Name", "Website", "State", "Industry",
        "Confidence", "Confidence Reason", "Resolved By", "Status",
    ])
    for row in rows:
        sheet.append(row)

    widths = {"A": 12, "B": 55, "C": 30, "D": 10, "E": 35, "F": 12, "G": 55, "H": 16, "I": 12}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Enrich approved sponsor list (Phase 2).")
    parser.add_argument("--input", type=Path, required=True, help="Input xlsx from Phase 1.")
    parser.add_argument("--output", type=Path, required=True, help="Output enriched xlsx.")
    parser.add_argument("--db", type=Path, default=Path("enrichment.db"), help="Checkpoint database path.")
    parser.add_argument("--limit", type=int, default=None, help="Only process this many new rows (test runs).")
    parser.add_argument("--skip-wikidata", action="store_true", help="Skip Tier 0, go straight to search.")
    parser.add_argument("--use-llm", action="store_true", help="Run the optional Gemini bonus pass on Low/None rows.")
    parser.add_argument("--llm-daily-cap", type=int, default=GEMINI_FREE_RPD, help="Cap for the bonus pass per run.")
    args = parser.parse_args()

    conn = init_db(args.db)
    companies = load_companies(args.input)
    print(f"Loaded {len(companies):,} companies from {args.input}")

    processed_this_run = 0
    consecutive_errors = 0

    for source_order, name in companies:
        if args.limit is not None and processed_this_run >= args.limit:
            print(f"Reached --limit {args.limit}; stopping (resumable).")
            break

        if already_done(conn, source_order):
            continue

        try:
            result = None
            if not args.skip_wikidata:
                result = wikidata_lookup(name)
                time.sleep(DDG_MIN_DELAY_SECONDS)

            if result is None:
                result = enrich_via_search(name)
                time.sleep(DDG_MIN_DELAY_SECONDS + random.uniform(0, 1.0))

            save_result(conn, source_order, name, result)
            processed_this_run += 1
            consecutive_errors = 0

            status_flag = "OK" if result["status"] == "done" else "unresolved"
            print(f"[{source_order}] {name[:60]:<60} -> {result.get('confidence', '?'):<6} ({status_flag})")

        except KeyboardInterrupt:
            print("\nInterrupted by user. Progress is saved - re-run the same command to resume.")
            break

        except Exception as exc:
            consecutive_errors += 1
            print(f"[{source_order}] {name[:60]:<60} -> ERROR: {exc}")
            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                print(f"{MAX_CONSECUTIVE_ERRORS} consecutive errors - stopping to check what's wrong "
                      "(likely DuckDuckGo throttling - wait a while before resuming).")
                break
            time.sleep(5)

    if args.use_llm:
        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        if not api_key:
            print("\n--use-llm was set but GEMINI_API_KEY is not in your environment - skipping bonus pass.")
        else:
            from google import genai
            client = genai.Client(api_key=api_key)
            llm_bonus_pass(conn, client, args.llm_daily_cap)

    conn.close()
    export_to_excel(args.db, args.output)
    print(f"\nExported current progress to {args.output}")
    print("Re-run this same command any time to continue enriching the remaining companies.")


if __name__ == "__main__":
    main()
